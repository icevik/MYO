#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import Workbook

class YerlesimApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Tercih Yerleştirme Uygulaması")
        self.geometry("800x600")

        # Seçilen Excel dosyasının yolu
        self.excel_path = None

        # Yerleştirme sonuçlarını saklayacağımız sözlük:
        # {
        #   "sirket_adi": [ "ogrenci_1", "ogrenci_2", ...],
        #   ...
        # }
        self.sonuc_yerlesimler = {}

        # Arayüz elemanlarını oluştur
        self.create_widgets()

        self.sirket_kontenjanlari = {}  # Yeni eklenecek

    def create_widgets(self):
        # Frame: Excel dosyası yükleme
        frame_upload = tk.Frame(self)
        frame_upload.pack(pady=10)

        btn_upload = tk.Button(frame_upload, text="Excel Yükle", command=self.excel_yukle)
        btn_upload.pack(side=tk.LEFT, padx=5)

        self.label_selected_file = tk.Label(frame_upload, text="Henüz bir dosya seçilmedi.")
        self.label_selected_file.pack(side=tk.LEFT, padx=5)

        # Frame: Yerleştirme işlemini başlatma
        frame_process = tk.Frame(self)
        frame_process.pack(pady=10)

        btn_process = tk.Button(frame_process, text="Yerleştir", command=self.yerlesim_islemi)
        btn_process.pack(side=tk.LEFT, padx=5)

        # Frame: Sonucu kaydetme
        frame_save = tk.Frame(self)
        frame_save.pack(pady=10)

        btn_save_excel = tk.Button(frame_save, text="Excel Olarak Kaydet", command=self.sonucu_excel_olarak_kaydet)
        btn_save_excel.pack(side=tk.LEFT, padx=5)

        # Text alanı: Sonuçları gösterme
        self.text_sonuc = tk.Text(self, height=20, width=90)
        self.text_sonuc.pack(pady=10)

    def excel_yukle(self):
        """Kullanıcıdan Excel dosyası seçmesini isteyen fonksiyon."""
        file_path = filedialog.askopenfilename(
            title="Excel Dosyası Seçin",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
        )
        if file_path:
            self.excel_path = file_path
            self.label_selected_file.config(text=f"Seçilen Dosya: {file_path}")
            self.kontenjan_penceresi_ac()

    def kontenjan_penceresi_ac(self):
        # Excel'den şirketleri oku
        wb = openpyxl.load_workbook(self.excel_path)
        sheet = wb.active
        
        sirketler = []
        sutun = 4
        while True:
            hucre_deger = sheet.cell(row=1, column=sutun).value
            if hucre_deger is None or hucre_deger == "":
                break
            sirketler.append(hucre_deger)
            sutun += 1
        
        # Popup pencere oluştur
        popup = tk.Toplevel(self)
        popup.title("Şirket Kontenjanları")
        popup.geometry("400x500")
        popup.grab_set()  # Pencereyi modal yap
        
        # Ana frame
        main_frame = tk.Frame(popup)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Başlık
        tk.Label(main_frame, text="Her şirket için kontenjan sayısını girin:", font=("Arial", 10, "bold")).pack(pady=10)
        
        # Scroll frame
        canvas = tk.Canvas(main_frame)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Kontenjan girişleri
        entries = {}
        for sirket in sirketler:
            frame = tk.Frame(scrollable_frame)
            frame.pack(fill="x", pady=5)
            
            tk.Label(frame, text=sirket, width=30, anchor="w").pack(side=tk.LEFT, padx=5)
            entry = tk.Entry(frame, width=10)
            entry.insert(0, "0")  # Varsayılan değer
            entry.pack(side=tk.LEFT, padx=5)
            entries[sirket] = entry
        
        def kaydet():
            try:
                hatali_girdi = False
                for sirket, entry in entries.items():
                    try:
                        kontenjan = int(entry.get())
                        if kontenjan < 0:
                            hatali_girdi = True
                            entry.config(bg="pink")
                        else:
                            entry.config(bg="white")
                    except ValueError:
                        hatali_girdi = True
                        entry.config(bg="pink")
                
                if hatali_girdi:
                    messagebox.showerror("Hata", "Lütfen tüm kontenjanlar için geçerli pozitif sayılar girin!")
                    return
                
                # Tüm girişler geçerliyse kaydet
                for sirket, entry in entries.items():
                    self.sirket_kontenjanlari[sirket] = int(entry.get())
                popup.destroy()
                
            except Exception as e:
                messagebox.showerror("Hata", f"Bir hata oluştu: {str(e)}")
        
        # Butonlar için frame
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Kaydet", command=kaydet, width=15).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="İptal", command=popup.destroy, width=15).pack(side=tk.LEFT, padx=5)
        
        # Scroll widget'larını yerleştir
        canvas.pack(side="left", fill="both", expand=True, pady=5)
        scrollbar.pack(side="right", fill="y")
        
        # Fare tekerleği ile scroll - sadece canvas üzerinde çalışsın
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<MouseWheel>", _on_mousewheel)  # Sadece canvas'a bind et
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)  # Frame'e de bind et

    def yerlesim_islemi(self):
        """
        Yerleştirme algoritmasını çalıştıran fonksiyon.
        Excel'den verileri okur, tercihlere göre şirketlere dağıtır ve
        sonuçları ekranda gösterir.
        """
        if not self.excel_path:
            messagebox.showwarning("Uyarı", "Lütfen önce bir Excel dosyası yükleyin.")
            return

        if not self.sirket_kontenjanlari:
            messagebox.showwarning("Uyarı", "Lütfen önce şirket kontenjanlarını belirleyin.")
            return

        try:
            # 1. Excel'i aç ve verileri oku
            wb = openpyxl.load_workbook(self.excel_path)
            sheet = wb.active  # Gerekirse wb["SayfaAdi"] şeklinde de yapabilirsiniz

            # Öğrencileri B sütunundan (B2'den itibaren) oku
            ogrenciler = []
            satir = 2
            while True:
                hucre_deger = sheet.cell(row=satir, column=2).value
                if hucre_deger is None or hucre_deger == "":
                    break
                ogrenciler.append(hucre_deger)
                satir += 1

            # Şirket isimlerini D1'den itibaren yatay olarak oku
            sirketler = []
            sutun = 4
            while True:
                hucre_deger = sheet.cell(row=1, column=sutun).value
                if hucre_deger is None or hucre_deger == "":
                    break
                sirketler.append(hucre_deger)
                sutun += 1

            M = len(ogrenciler)
            N = len(sirketler)

            if M == 0 or N == 0:
                messagebox.showinfo("Bilgi", "Öğrenci veya şirket bulunamadı. Excel formatını kontrol edin.")
                return

            # 2. Öğrencilerin tercih bilgilerini oku
            tercih_dict = {}
            for i, ogrenci in enumerate(ogrenciler, start=2):
                ogrenci_adi = ogrenci
                tercih_dict[ogrenci_adi] = {}
                for j, sirket in enumerate(sirketler, start=4):
                    tercih_sirasi = sheet.cell(row=i, column=j).value
                    if tercih_sirasi is not None and str(tercih_sirasi).strip() != "":
                        try:
                            tercih_sirasi = int(tercih_sirasi)
                        except ValueError:
                            # Hücrede sayı dışında bir şey varsa burada karar verilebilir
                            # Şimdilik geçiyoruz
                            pass
                        tercih_dict[ogrenci_adi][sirket] = tercih_sirasi

            # 3. Kapasiteleri belirle
            kapasiteler = self.sirket_kontenjanlari.copy()

            # 4. Yerleştirme
            # sonuc_yerlesimler dict'inde tutacağız
            self.sonuc_yerlesimler = {s: [] for s in sirketler}

            for ogrenci_adi in ogrenciler:
                # O öğrencinin tercihlerini al
                tercih_listesi = list(tercih_dict[ogrenci_adi].items())  # [(sirket, sira), ...]
                # Tercih sırasına göre sırala (1 en yüksek)
                tercih_listesi.sort(key=lambda x: x[1])

                # Kapasite müsaitse her tercihe yerleştir
                for (sirket_adi, sira) in tercih_listesi:
                    if kapasiteler[sirket_adi] > 0:
                        self.sonuc_yerlesimler[sirket_adi].append(ogrenci_adi)
                        kapasiteler[sirket_adi] -= 1
                    # "Bir öğrenci birden fazla şirkete gidebilir" dediğimiz için
                    # diğer tercihleri de kontrol ediyoruz (kapasite uygun oldukça).

            # 5. Sonuçları ekrana yaz
            self.goruntule_sonuc()

            # Tercih yapmayan ve yerleşemeyen öğrencileri takip etmek için listeler
            tercih_yapmayan = []
            yerlesemeyen = []
            
            # Tercih kontrolü
            for ogrenci_adi in ogrenciler:
                if not tercih_dict[ogrenci_adi]:  # Hiç tercih yoksa
                    tercih_yapmayan.append(ogrenci_adi)
                    continue
                
            # Yerleştirme sonrası kontrol
            yerlesen_ogrenciler = set()
            for sirket, ogrenciler in self.sonuc_yerlesimler.items():
                yerlesen_ogrenciler.update(ogrenciler)
            
            # Yerleşemeyen öğrencileri bul
            for ogrenci in ogrenciler:
                if ogrenci not in yerlesen_ogrenciler and ogrenci not in tercih_yapmayan:
                    yerlesemeyen.append(ogrenci)
            
            # Sonuçlara ekle
            self.sonuc_yerlesimler["TERCİH YAPMAYAN"] = tercih_yapmayan
            self.sonuc_yerlesimler["YERLEŞEMEYEN"] = yerlesemeyen

        except Exception as e:
            messagebox.showerror("Hata", f"Bir hata oluştu:\n{e}")

    def goruntule_sonuc(self):
        """Yerleştirme sonuçlarını Text alanına yazan fonksiyon."""
        self.text_sonuc.delete("1.0", tk.END)  # Önce temizleyelim

        if not self.sonuc_yerlesimler:
            self.text_sonuc.insert(tk.END, "Henüz bir sonuç yok.\n")
            return

        # Önce normal şirketleri göster
        for sirket, ogrenci_listesi in self.sonuc_yerlesimler.items():
            if sirket not in ["TERCİH YAPMAYAN", "YERLEŞEMEYEN"]:
                self.text_sonuc.insert(tk.END, f"\nŞirket: {sirket}\n")
                if not ogrenci_listesi:
                    self.text_sonuc.insert(tk.END, "  Yerleşen öğrenci yok.\n")
                else:
                    for ogr in ogrenci_listesi:
                        self.text_sonuc.insert(tk.END, f"  - {ogr}\n")
        
        # Özel durumları göster
        ozel_durumlar = ["TERCİH YAPMAYAN", "YERLEŞEMEYEN"]
        for durum in ozel_durumlar:
            if durum in self.sonuc_yerlesimler:
                self.text_sonuc.insert(tk.END, f"\n{durum} ÖĞRENCİLER:\n")
                ogrenciler = self.sonuc_yerlesimler[durum]
                if not ogrenciler:
                    self.text_sonuc.insert(tk.END, "  Öğrenci yok.\n")
                else:
                    for ogr in ogrenciler:
                        self.text_sonuc.insert(tk.END, f"  - {ogr}\n")

    def sonucu_excel_olarak_kaydet(self):
        """
        Yerleştirme sonuçlarını yeni bir Excel dosyasına yazıp
        kullanıcıya kaydetme konumu seçtiren fonksiyon.
        """
        if not self.sonuc_yerlesimler:
            messagebox.showinfo("Bilgi", "Önce yerleştirme yapın, sonuç oluşturun.")
            return

        # Kaydetme konumu sor
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx"), ("Tüm Dosyalar", "*.*")]
        )
        if not save_path:
            return  # Kullanıcı iptal etmiş

        try:
            yeni_wb = Workbook()
            yeni_sheet = yeni_wb.active
            yeni_sheet.title = "YerlesimSonuclari"

            # Şirketleri üst satıra yaz
            sirketler = list(self.sonuc_yerlesimler.keys())
            for col_idx, sirket in enumerate(sirketler, start=1):
                yeni_sheet.cell(row=1, column=col_idx, value=sirket)

            # En çok öğrencisi olan şirketin uzunluğunu bul
            max_len = max(len(self.sonuc_yerlesimler[s]) for s in sirketler)

            # Her satırda ilgili şirketin öğrenci listesinden indexleyerek yaz
            for row_idx in range(2, max_len + 2):
                for col_idx, sirket in enumerate(sirketler, start=1):
                    ogrenci_listesi = self.sonuc_yerlesimler[sirket]
                    index_ = row_idx - 2
                    if index_ < len(ogrenci_listesi):
                        yeni_sheet.cell(row=row_idx, column=col_idx, value=ogrenci_listesi[index_])
                    else:
                        yeni_sheet.cell(row=row_idx, column=col_idx, value="")

            # Özel sütunları en sona ekle
            ozel_sutunlar = ["TERCİH YAPMAYAN", "YERLEŞEMEYEN"]
            for sutun in ozel_sutunlar:
                if sutun in self.sonuc_yerlesimler:
                    sirketler.append(sutun)

            yeni_wb.save(save_path)
            messagebox.showinfo("Bilgi", f"Sonuçlar kaydedildi: {save_path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kayıt sırasında bir hata oluştu:\n{e}")


if __name__ == "__main__":
    app = YerlesimApp()
    app.mainloop()
